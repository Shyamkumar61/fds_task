from celery import shared_task
from openpyxl import load_workbook
from .models import Product

@shared_task(bind=True)
def process_bulk_upload(self, file_name, file_content):
    wb = load_workbook(filename=file_name, read_only=True)
    ws = wb.active

    total_rows = ws.max_row

    for row in ws.iter_rows(min_row=2, values_only=True):
        Product.objects.create(
            name=row[0],
            description=row[1],
            price=row[2]
        )
        self.update_state(state='PROGRESS', meta={'current': ws.max_row - row[0], 'total': total_rows})
    return {'current': total_rows, 'total': total_rows, 'status': 'Task completed!'}